from openpyxl import load_workbook


class ExcelReader:
    def __init__(self, file_path: str, column_no: int) -> None:
        self._celebrants: list = []
        self._column_no: int = column_no
        self._workbook = load_workbook(file_path)
        self._active_sheet = None

    def set_active_sheet(self, sheet_name: str):
        self._active_sheet = self._workbook[sheet_name]

    def get_active_sheet(self):
        return self._active_sheet

    def get_celebrants_from_file(self, month, day):
        max_row = self._active_sheet.max_row
        for row_num in range(2, max_row+1):
            birthday = self._active_sheet.cell(row=row_num, column=self._column_no)
        
            if birthday.value.day == day and birthday.value.month == month:
                first_name = self._active_sheet.cell(row=row_num, column=1).value
                last_name = self._active_sheet.cell(row=row_num, column=2).value
                full_name = first_name + " " + last_name
                self._celebrants.append(full_name)

        return self._celebrants
